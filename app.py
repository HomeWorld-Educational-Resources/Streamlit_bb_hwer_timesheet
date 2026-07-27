
import json
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook  

try:
    from tkcalendar import DateEntry  
    HAS_TKCALENDAR = True
except Exception:
    DateEntry = None
    HAS_TKCALENDAR = False

APP_DIR = Path(__file__).resolve().parent
DEFAULT_TEMPLATE = APP_DIR / 'customer_timesheet_gui_template.xlsx'
CUSTOMERS_JSON = APP_DIR / 'customers_cache.json'

ACTIVITY_ROWS = [22, 23, 24, 25, 26]
MAX_PAYMENT_ROWS = 30
PAYMENT_ROWS = list(range(32, 32 + MAX_PAYMENT_ROWS))
PAYMENT_DATE_FORMAT = '%b-%d-%Y'   # Jan-26-2026
WEEK_DATE_FORMAT = '%Y-%m-%d'


class ScrollableFrame(ttk.Frame):
    def __init__(self, master, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        canvas = tk.Canvas(self, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient='vertical', command=canvas.yview)
        self.inner = ttk.Frame(canvas)
        self.inner.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=self.inner, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')

        canvas.bind_all('<MouseWheel>', _on_mousewheel)
        self.canvas = canvas


class TimesheetApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Customer Timesheet Builder')
        self.root.geometry('1220x880')

        self.template_path = tk.StringVar(value=str(DEFAULT_TEMPLATE))
        self.output_dir = tk.StringVar(value=str(APP_DIR / 'output'))
        Path(self.output_dir.get()).mkdir(parents=True, exist_ok=True)

        self.customers = []
        self.customer_lookup = {}
        self.activity_vars = []
        self.payment_vars = []
        self.payment_date_widgets = []

        self._build_ui()
        self.load_customers()
        self.load_defaults_from_template()

    def _build_ui(self):
        top = ttk.Frame(self.root, padding=10)
        top.pack(fill='x')

        ttk.Label(top, text='Template workbook').grid(row=0, column=0, sticky='w')
        ttk.Entry(top, textvariable=self.template_path, width=92).grid(row=0, column=1, sticky='ew', padx=6)
        ttk.Button(top, text='Browse', command=self.choose_template).grid(row=0, column=2)

        ttk.Label(top, text='Output folder').grid(row=1, column=0, sticky='w', pady=(6, 0))
        ttk.Entry(top, textvariable=self.output_dir, width=92).grid(row=1, column=1, sticky='ew', padx=6, pady=(6, 0))
        ttk.Button(top, text='Browse', command=self.choose_output_dir).grid(row=1, column=2, pady=(6, 0))
        top.columnconfigure(1, weight=1)

        body = ScrollableFrame(self.root)
        body.pack(fill='both', expand=True, padx=10, pady=(0, 10))
        main = body.inner

        customer_frame = ttk.LabelFrame(main, text='Customer + Contract', padding=10)
        customer_frame.pack(fill='x', pady=(0, 10))

        self.selected_customer = tk.StringVar()
        self.new_customer_name = tk.StringVar()
        self.company_project = tk.StringVar()
        self.default_rate = tk.StringVar()
        self.max_spend = tk.StringVar()
        self.contract_note = tk.StringVar()
        self.footnote = tk.StringVar()

        ttk.Label(customer_frame, text='Existing customer name').grid(row=0, column=0, sticky='w')
        self.customer_combo = ttk.Combobox(customer_frame, textvariable=self.selected_customer, width=32, state='readonly')
        self.customer_combo.grid(row=0, column=1, sticky='w', padx=6)
        self.customer_combo.bind('<<ComboboxSelected>>', self.on_customer_selected)

        ttk.Label(customer_frame, text='Create new customer name').grid(row=0, column=2, sticky='w', padx=(20, 0))
        ttk.Entry(customer_frame, textvariable=self.new_customer_name, width=30).grid(row=0, column=3, sticky='w', padx=6)

        ttk.Label(customer_frame, text='Company / project').grid(row=1, column=0, sticky='w', pady=(8, 0))
        ttk.Entry(customer_frame, textvariable=self.company_project, width=36).grid(row=1, column=1, sticky='w', padx=6, pady=(8, 0))

        ttk.Label(customer_frame, text='Hourly rate').grid(row=1, column=2, sticky='w', padx=(20, 0), pady=(8, 0))
        ttk.Entry(customer_frame, textvariable=self.default_rate, width=14).grid(row=1, column=3, sticky='w', padx=6, pady=(8, 0))

        ttk.Label(customer_frame, text='Max contract spend').grid(row=2, column=0, sticky='w', pady=(8, 0))
        ttk.Entry(customer_frame, textvariable=self.max_spend, width=20).grid(row=2, column=1, sticky='w', padx=6, pady=(8, 0))

        ttk.Label(customer_frame, text='Contract note').grid(row=2, column=2, sticky='w', padx=(20, 0), pady=(8, 0))
        ttk.Entry(customer_frame, textvariable=self.contract_note, width=42).grid(row=2, column=3, sticky='w', padx=6, pady=(8, 0))

        ttk.Label(customer_frame, text='Footnote').grid(row=3, column=0, sticky='w', pady=(8, 0))
        ttk.Entry(customer_frame, textvariable=self.footnote, width=86).grid(row=3, column=1, columnspan=3, sticky='ew', padx=6, pady=(8, 0))

        btns = ttk.Frame(customer_frame)
        btns.grid(row=4, column=0, columnspan=4, sticky='w', pady=(10, 0))
        ttk.Button(btns, text='Save / Update Customer', command=self.save_customer).pack(side='left')
        ttk.Button(btns, text='Reload Customers', command=self.load_customers).pack(side='left', padx=8)

        week_frame = ttk.LabelFrame(main, text='Week Setup', padding=10)
        week_frame.pack(fill='x', pady=(0, 10))

        self.week_start = tk.StringVar()
        self.week_number = tk.StringVar()
        self.hourly_rate = tk.StringVar()
        self.prior_balance = tk.StringVar()
        self.max_contract_spend_override = tk.StringVar()

        ttk.Label(week_frame, text='Week start date (YYYY-MM-DD)').grid(row=0, column=0, sticky='w')
        ttk.Entry(week_frame, textvariable=self.week_start, width=18).grid(row=0, column=1, sticky='w', padx=6)
        ttk.Label(week_frame, text='Week number').grid(row=0, column=2, sticky='w', padx=(20, 0))
        ttk.Entry(week_frame, textvariable=self.week_number, width=10).grid(row=0, column=3, sticky='w', padx=6)
        ttk.Label(week_frame, text='Hourly rate override').grid(row=1, column=0, sticky='w', pady=(8, 0))
        ttk.Entry(week_frame, textvariable=self.hourly_rate, width=18).grid(row=1, column=1, sticky='w', padx=6, pady=(8, 0))
        ttk.Label(week_frame, text='Prior balance').grid(row=1, column=2, sticky='w', padx=(20, 0), pady=(8, 0))
        ttk.Entry(week_frame, textvariable=self.prior_balance, width=18).grid(row=1, column=3, sticky='w', padx=6, pady=(8, 0))
        ttk.Label(week_frame, text='Max contract spend override').grid(row=2, column=0, sticky='w', pady=(8, 0))
        ttk.Entry(week_frame, textvariable=self.max_contract_spend_override, width=18).grid(row=2, column=1, sticky='w', padx=6, pady=(8, 0))

        activity_frame = ttk.LabelFrame(main, text='Activities (5 rows)', padding=10)
        activity_frame.pack(fill='x', pady=(0, 10))
        headers = ['Activity', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Output description']
        for c, h in enumerate(headers):
            ttk.Label(activity_frame, text=h).grid(row=0, column=c, sticky='w', padx=4, pady=(0, 6))

        for idx in range(5):
            row_vars = {
                'activity': tk.StringVar(),
                'mon': tk.StringVar(value='0'),
                'tue': tk.StringVar(value='0'),
                'wed': tk.StringVar(value='0'),
                'thu': tk.StringVar(value='0'),
                'fri': tk.StringVar(value='0'),
                'output': tk.StringVar(),
            }
            self.activity_vars.append(row_vars)
            ttk.Entry(activity_frame, textvariable=row_vars['activity'], width=36).grid(row=idx + 1, column=0, sticky='ew', padx=4, pady=2)
            for c, day in enumerate(['mon', 'tue', 'wed', 'thu', 'fri'], start=1):
                ttk.Entry(activity_frame, textvariable=row_vars[day], width=8).grid(row=idx + 1, column=c, sticky='ew', padx=4, pady=2)
            ttk.Entry(activity_frame, textvariable=row_vars['output'], width=60).grid(row=idx + 1, column=6, sticky='ew', padx=4, pady=2)

        payment_frame = ttk.LabelFrame(main, text=f'Payment Records (up to {MAX_PAYMENT_ROWS} rows)', padding=10)
        payment_frame.pack(fill='x', pady=(0, 10))
        for c, h in enumerate(['Amount', 'Payment date (Jan-26-2026)', 'Notes']):
            ttk.Label(payment_frame, text=h).grid(row=0, column=c, sticky='w', padx=4, pady=(0, 6))

        for idx in range(MAX_PAYMENT_ROWS):
            row_vars = {
                'amount': tk.StringVar(),
                'date': tk.StringVar(),
                'notes': tk.StringVar(),
            }
            self.payment_vars.append(row_vars)
            ttk.Entry(payment_frame, textvariable=row_vars['amount'], width=14).grid(row=idx + 1, column=0, sticky='w', padx=4, pady=2)

            if HAS_TKCALENDAR:
                date_widget = DateEntry(
                    payment_frame,
                    width=16,
                    date_pattern='mm-dd-yyyy',
                    textvariable=row_vars['date'],
                )
                date_widget.grid(row=idx + 1, column=1, sticky='w', padx=4, pady=2)
            else:
                date_widget = ttk.Entry(payment_frame, textvariable=row_vars['date'], width=18)
                date_widget.grid(row=idx + 1, column=1, sticky='w', padx=4, pady=2)
            self.payment_date_widgets.append(date_widget)

            ttk.Entry(payment_frame, textvariable=row_vars['notes'], width=70).grid(row=idx + 1, column=2, sticky='ew', padx=4, pady=2)

        actions = ttk.Frame(main)
        actions.pack(fill='x', pady=(4, 12))
        ttk.Button(actions, text='Load From Template', command=self.load_defaults_from_template).pack(side='left')
        ttk.Button(actions, text='Generate Filled Workbook', command=self.generate_workbook).pack(side='left', padx=10)
        ttk.Button(actions, text='Reset Form', command=self.reset_form).pack(side='left')

        self.status = tk.StringVar(value='Ready.')
        ttk.Label(main, textvariable=self.status).pack(anchor='w')

    def choose_template(self):
        path = filedialog.askopenfilename(title='Choose template workbook', filetypes=[('Excel files', '*.xlsx')])
        if path:
            self.template_path.set(path)
            self.load_customers()
            self.load_defaults_from_template()

    def choose_output_dir(self):
        path = filedialog.askdirectory(title='Choose output folder')
        if path:
            self.output_dir.set(path)
            Path(path).mkdir(parents=True, exist_ok=True)

    def _safe_float(self, value, allow_blank=True):
        text = str(value).strip()
        if not text:
            return None if allow_blank else 0.0
        try:
            return float(text)
        except ValueError:
            raise ValueError(f'Invalid number: {value}')

    def _safe_date(self, value, allow_blank=True, kind='week'):
        text = str(value).strip()
        if not text:
            return None if allow_blank else None

        formats = [
            '%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d',
            '%b-%d-%Y', '%B-%d-%Y',
            '%m-%d-%Y', '%m/%d/%y', '%m-%d-%y',
        ]
        for fmt in formats:
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue

        if kind == 'payment':
            raise ValueError(f'Invalid payment date: {value}. Use Jan-26-2026 or pick from the calendar.')
        raise ValueError(f'Invalid date: {value}. Use YYYY-MM-DD.')

    def _format_payment_date(self, value):
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.strftime(PAYMENT_DATE_FORMAT)
        if hasattr(value, 'strftime'):
            return value.strftime(PAYMENT_DATE_FORMAT)
        text = str(value).strip()
        if not text or text.startswith('='):
            return ''
        parsed = self._safe_date(text, allow_blank=True, kind='payment')
        return parsed.strftime(PAYMENT_DATE_FORMAT) if parsed else text

    def load_customers(self):
        self.customers = []
        self.customer_lookup = {}
        path = Path(self.template_path.get())
        if not path.exists():
            self.customer_combo['values'] = []
            self.status.set('Template workbook not found yet.')
            return

        try:
            wb = load_workbook(path)
            ws = wb['Customers']
            for row in range(2, ws.max_row + 1):
                name = ws[f'A{row}'].value
                if not name or str(name).strip().lower() == 'how to use':
                    continue
                data = {
                    'name': str(name).strip(),
                    'company_project': ws[f'B{row}'].value or '',
                    'default_rate': ws[f'C{row}'].value if ws[f'C{row}'].value is not None else '',
                    'max_contract_spend': ws[f'D{row}'].value if ws[f'D{row}'].value is not None else '',
                    'notes': ws[f'E{row}'].value or '',
                }
                self.customers.append(data)
                self.customer_lookup[data['name']] = data

            names = [c['name'] for c in self.customers]
            self.customer_combo['values'] = names
            current_name = self.selected_customer.get().strip()
            if current_name and current_name in self.customer_lookup:
                self.selected_customer.set(current_name)
                self.on_customer_selected()
            elif self.customers:
                self.selected_customer.set(self.customers[0]['name'])
                self.on_customer_selected()

            with open(CUSTOMERS_JSON, 'w', encoding='utf-8') as f:
                json.dump(self.customers, f, indent=2, default=str)
            self.status.set(f'Loaded {len(self.customers)} customer(s).')
        except Exception as e:
            messagebox.showerror('Load customers failed', str(e))

    def on_customer_selected(self, event=None):
        name = self.selected_customer.get().strip()
        data = self.customer_lookup.get(name)
        if not data:
            return
        self.new_customer_name.set(data['name'])
        self.company_project.set(str(data.get('company_project', '')))
        self.default_rate.set(str(data.get('default_rate', '')))
        self.max_spend.set(str(data.get('max_contract_spend', '')))

    def save_customer(self):
        template = Path(self.template_path.get())
        if not template.exists():
            messagebox.showerror('Missing template', 'Please choose a valid template workbook first.')
            return

        raw_name = self.new_customer_name.get().strip() or self.selected_customer.get().strip()
        if not raw_name:
            messagebox.showerror('Missing customer name', 'Enter a customer name.')
            return

        normalized_name = raw_name.casefold()

        try:
            wb = load_workbook(template)
            ws = wb['Customers']
            target_row = None
            exact_existing_name = None

            for row in range(2, max(ws.max_row, 2) + 1):
                existing = ws[f'A{row}'].value
                if not existing:
                    continue
                existing_text = str(existing).strip()
                if existing_text.lower() == 'how to use':
                    continue
                if existing_text.casefold() == normalized_name:
                    target_row = row
                    exact_existing_name = existing_text
                    break

            if target_row is None:
                target_row = max(ws.max_row + 1, 2)
                while str(ws[f'A{target_row}'].value or '').strip().lower() == 'how to use':
                    target_row += 1
                final_name = raw_name
            else:
                final_name = exact_existing_name

            ws[f'A{target_row}'] = final_name
            ws[f'B{target_row}'] = self.company_project.get().strip()
            ws[f'C{target_row}'] = self._safe_float(self.default_rate.get())
            ws[f'D{target_row}'] = self._safe_float(self.max_spend.get())
            ws[f'E{target_row}'] = self.contract_note.get().strip()
            wb.save(template)

            self.load_customers()
            self.selected_customer.set(final_name)
            self.new_customer_name.set(final_name)
            self.on_customer_selected()
            messagebox.showinfo('Customer saved', f'Customer "{final_name}" has been saved to the Customers sheet.')
        except Exception as e:
            messagebox.showerror('Save customer failed', str(e))

    def load_defaults_from_template(self):
        template = Path(self.template_path.get())
        if not template.exists():
            return
        try:
            wb = load_workbook(template)
            ws = wb['Control Panel']
            if ws['B4'].value:
                self.selected_customer.set(str(ws['B4'].value))
            self.week_start.set(self._date_to_string(ws['B12'].value))
            self.week_number.set(str(ws['B13'].value or ''))
            self.hourly_rate.set(self._formula_or_value_to_string(ws['B14'].value))
            self.prior_balance.set(self._formula_or_value_to_string(ws['B15'].value))
            self.max_contract_spend_override.set(self._formula_or_value_to_string(ws['B16'].value))
            self.contract_note.set(self._formula_or_value_to_string(ws['B8'].value))
            self.footnote.set(self._formula_or_value_to_string(ws['B9'].value))

            for i, row in enumerate(ACTIVITY_ROWS):
                self.activity_vars[i]['activity'].set(self._formula_or_value_to_string(ws[f'A{row}'].value))
                self.activity_vars[i]['mon'].set(self._formula_or_value_to_string(ws[f'B{row}'].value, default='0'))
                self.activity_vars[i]['tue'].set(self._formula_or_value_to_string(ws[f'C{row}'].value, default='0'))
                self.activity_vars[i]['wed'].set(self._formula_or_value_to_string(ws[f'D{row}'].value, default='0'))
                self.activity_vars[i]['thu'].set(self._formula_or_value_to_string(ws[f'E{row}'].value, default='0'))
                self.activity_vars[i]['fri'].set(self._formula_or_value_to_string(ws[f'F{row}'].value, default='0'))
                self.activity_vars[i]['output'].set(self._formula_or_value_to_string(ws[f'H{row}'].value))

            for i, row in enumerate(PAYMENT_ROWS):
                self.payment_vars[i]['amount'].set(self._formula_or_value_to_string(ws[f'A{row}'].value))
                self.payment_vars[i]['date'].set(self._format_payment_date(ws[f'B{row}'].value))
                self.payment_vars[i]['notes'].set(self._formula_or_value_to_string(ws[f'C{row}'].value))

            if self.selected_customer.get() in self.customer_lookup:
                self.on_customer_selected()
            self.status.set('Loaded defaults from template.')
        except Exception as e:
            messagebox.showerror('Load template failed', str(e))

    def _date_to_string(self, value):
        if isinstance(value, datetime):
            return value.strftime(WEEK_DATE_FORMAT)
        if hasattr(value, 'strftime'):
            return value.strftime(WEEK_DATE_FORMAT)
        if value is None:
            return ''
        text = str(value)
        if text.startswith('='):
            return ''
        return text

    def _formula_or_value_to_string(self, value, default=''):
        if value is None:
            return default
        text = str(value)
        if text.startswith('='):
            return default
        return text

    def reset_form(self):
        self.new_customer_name.set('')
        self.company_project.set('')
        self.default_rate.set('')
        self.max_spend.set('')
        self.contract_note.set('')
        self.footnote.set('')
        self.week_start.set('')
        self.week_number.set('')
        self.hourly_rate.set('')
        self.prior_balance.set('')
        self.max_contract_spend_override.set('')
        for row in self.activity_vars:
            row['activity'].set('')
            row['mon'].set('0')
            row['tue'].set('0')
            row['wed'].set('0')
            row['thu'].set('0')
            row['fri'].set('0')
            row['output'].set('')
        for row in self.payment_vars:
            row['amount'].set('')
            row['date'].set('')
            row['notes'].set('')
        self.status.set('Form reset.')

    def generate_workbook(self):
        template = Path(self.template_path.get())
        if not template.exists():
            messagebox.showerror('Missing template', 'Please choose a valid template workbook.')
            return
        try:
            wb = load_workbook(template)
            ws = wb['Control Panel']

            customer_name = self.new_customer_name.get().strip() or self.selected_customer.get().strip()
            if not customer_name:
                raise ValueError('Please select or enter a customer name.')

            ws['B4'] = customer_name
            ws['B6'] = self.company_project.get().strip() or None
            ws['B8'] = self.contract_note.get().strip() or None
            ws['B9'] = self.footnote.get().strip() or None

            week_start = self._safe_date(self.week_start.get(), allow_blank=False, kind='week')
            ws['B12'] = week_start
            ws['B13'] = int(float(self.week_number.get().strip() or 0))

            hourly_rate = self._safe_float(self.hourly_rate.get())
            prior_balance = self._safe_float(self.prior_balance.get(), allow_blank=False)
            max_contract = self._safe_float(self.max_contract_spend_override.get())

            if hourly_rate is not None:
                ws['B14'] = hourly_rate
            ws['B15'] = prior_balance
            if max_contract is not None:
                ws['B16'] = max_contract

            for i, row in enumerate(ACTIVITY_ROWS):
                data = self.activity_vars[i]
                ws[f'A{row}'] = data['activity'].get().strip() or None
                ws[f'B{row}'] = self._safe_float(data['mon'].get(), allow_blank=False)
                ws[f'C{row}'] = self._safe_float(data['tue'].get(), allow_blank=False)
                ws[f'D{row}'] = self._safe_float(data['wed'].get(), allow_blank=False)
                ws[f'E{row}'] = self._safe_float(data['thu'].get(), allow_blank=False)
                ws[f'F{row}'] = self._safe_float(data['fri'].get(), allow_blank=False)
                ws[f'H{row}'] = data['output'].get().strip() or None

            for i, row in enumerate(PAYMENT_ROWS):
                data = self.payment_vars[i]
                amount = self._safe_float(data['amount'].get())
                note = data['notes'].get().strip()
                raw_date = data['date'].get().strip()
                formatted_date = None
                if raw_date:
                    pay_date = self._safe_date(raw_date, allow_blank=True, kind='payment')
                    if pay_date:
                        formatted_date = pay_date.strftime(PAYMENT_DATE_FORMAT)
                ws[f'A{row}'] = amount
                ws[f'B{row}'] = formatted_date or None
                ws[f'C{row}'] = note or None

            from openpyxl.workbook.properties import CalcProperties
            if wb.calculation is None:
                wb.calculation = CalcProperties(calcMode='auto', fullCalcOnLoad=True, forceFullCalc=True)
            else:
                wb.calculation.fullCalcOnLoad = True
                wb.calculation.forceFullCalc = True
                wb.calculation.calcMode = 'auto'

            safe_customer = ''.join(ch if ch.isalnum() or ch in (' ', '-', '_') else '_' for ch in customer_name).strip() or 'Customer'
            year = week_start.strftime('%Y')
            month = week_start.strftime('%b')
            base_name = f'{safe_customer} - {year} {month} Week {ws["B13"].value} timesheet'
            output_path = Path(self.output_dir.get()) / f'{base_name}.xlsx'
            Path(self.output_dir.get()).mkdir(parents=True, exist_ok=True)
            counter = 2
            while output_path.exists():
                output_path = Path(self.output_dir.get()) / f'{base_name} ({counter}).xlsx'
                counter += 1
            wb.save(output_path)

            self.status.set(f'Workbook generated: {output_path}')
            messagebox.showinfo('Success', f'Workbook created:\n{output_path}')
        except Exception as e:
            messagebox.showerror('Generate workbook failed', str(e))


def ensure_template_present():
    if DEFAULT_TEMPLATE.exists():
        return
    src = Path('/mnt/data/customer_timesheet_gui_template.xlsx')
    if src.exists():
        shutil.copy2(src, DEFAULT_TEMPLATE)


if __name__ == '__main__':
    ensure_template_present()
    root = tk.Tk()
    style = ttk.Style(root)
    try:
        style.theme_use('clam')
    except Exception:
        pass
    app = TimesheetApp(root)
    root.mainloop()
